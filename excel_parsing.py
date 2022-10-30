import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#wb = WorkBook()
wb = load_workbook('Wyndor.xlsx')

### modifying speciecic workbook sheets
ws = wb.active ## set object to active worksheet
print(ws)
print(ws['A1'].value)

a2 = ws['a2']
print(a2)
ws.title = 'data' ## change data
## inserting single points of data i
ws['A1'] = 24
ws['A2'] = 44

## ws.cell to access by row and column notation
d = ws.cell(row=4, column=2, value=10)
print(d)


## creating new worksheet
ws1 = wb.create_sheet("Mysheet") ## insert in first sheet
ws1.title = 'Sheet One'
ws1.sheet_properties.tabColor = "1072BA"
ws2 = wb.create_sheet("Mysheet_first position", 0) ## insert in second sheet
ws2.title = 'sheet two'
ws2.sheet_properties.tabColor = "1072BA"

print('ws1', ws1)
print('ws2', ws2)

print(wb.sheetnames)

## to acces specific sheet
print('accessing specific sheet')
print(wb[ws1])


for sheet in wb:
    print(sheet.title)

## to access cells
for row in range(1, 9):
    for col in range(1, 15):
        char = get_column_letter(col)
        print(ws[char + str(row)])


# create 100 x 100 cell . this scrollls trhough the cells, rather than access them, which stores in memory
for x in range(1,101):
    for y in range(1,101):
        ws.cell(row=x, column=y)
        print(ws.cell)


## writing data onto excel sheetnames
ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")
ws.deleterow()
## append multiple lines to worksheet
ws.append(['Adel', 'was', 'here'])
print(wb.sheetnames) ## print all sheets




### create copys of worksheeet ###
source = wb.active
target = wb.copy_worksheet(source)
print(source)
source_ws1 = ws1
target_ws1 = wb.copy_worksheet(source_ws1)
print(target_ws1)




